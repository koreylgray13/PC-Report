import os
import glob
import pandas as pd
import openpyxl
import xlwings as xw
from tkinter import messagebox
from datetime import datetime as dt
from functools import cache


# Define Path Variables
dataPath = "A:"
destinationDP = "C:\\Users\\kgray\\Herff Jones\\HJ Indy Data - Documents\\Production Control.xlsx"
destinationFP = "C:\\Users\\kgray\\Herff Jones\\HJ Indy Data - Documents\\Production Control_FP.xlsx"


# Clear Reports
@cache
def clearDP():
    # Identify & Open Destination File
    destFileDP = openpyxl.load_workbook(destinationDP)
    destSheetDP = destFileDP['ALL JOBS']

    # Loop Through Cells & Clear Previous Data
    for a in destSheetDP['F2':'BF3000']:
        for cell in a:
            cell.value = None

    # Save Workbook After Clearing Data
    destFileDP.save(destinationDP)
    messagebox.showinfo("Status", "Finished.")


@cache
def clearFP():
    # Identify & Open Destination File
    destFileFP = openpyxl.load_workbook(destinationFP)
    destSheetFP = destFileFP['All Jobs_FP']

    # Loop Through Cells & Clear Previous Data
    for a in destSheetFP['E2':'BF3000']:
        for cell in a:
            cell.value = None

    # Save Workbook After Clearing Data
    destFileFP.save(destinationFP)
    messagebox.showinfo("Status", "Finished.")


# Locate Files
@cache
def findFileDP():
    global targetFileDP
    # Change Directories & Identify Target File
    os.chdir(dataPath)
    # Find Highest Numbered Report
    targetFileDP = max(glob.glob("DP_All_Jobs_Report*.csv"))
    return targetFileDP


@cache
def findFileFP():
    global targetFileFP
    # Identify Target File
    os.chdir(dataPath)
    # Find Highest Numbered Report
    targetFileFP = max(glob.glob("FP_All_Jobs_Report*.csv"))
    return targetFileFP


# File Properties
@cache
def statsDP():
    global targetFileDP, stats

    # Change To Source Data Directory
    os.chdir(dataPath)

    # Define File
    file = findFileDP()

    # File Properties
    stats = (os.stat(file))

    # Reference Modified Time
    modified_time = (os.stat(file).st_mtime)
    modified_time = dt.fromtimestamp(modified_time).strftime("%I:%M:%S %p - %A")

    # Reference Access Time
    access_time = (os.stat(file).st_atime)
    access_time = dt.fromtimestamp(access_time).strftime("%I:%M:%S %p - %A")

    # Reference Created Time
    created_time = (os.stat(file).st_ctime)
    created_time = dt.fromtimestamp(created_time).strftime("%I:%M:%S %p - %A")
    # Format Response
    stats = ("\nCreated: \t\t{}\nModified: \t{}\nAccessed: \t{}".format(created_time, modified_time, access_time))
    return stats


@cache
def statsFP():
    global targetFileFP, stats

    # Change To Source Data Directory
    os.chdir(dataPath)

    # Define File
    file = findFileDP()

    # File Properties
    stats = (os.stat(file))

    # Reference Modified Time
    modified_time = (os.stat(file).st_mtime)
    modified_time = dt.fromtimestamp(modified_time).strftime("%I:%M:%S %p - %A")

    # Reference Access Time
    access_time = (os.stat(file).st_atime)
    access_time = dt.fromtimestamp(access_time).strftime("%I:%M:%S %p - %A")

    # Reference Created Time
    created_time = (os.stat(file).st_ctime)
    created_time = dt.fromtimestamp(created_time).strftime("%I:%M:%S %p - %A")

    stats = ("\nCreated: \t\t{}\nModified: \t{}\nAccessed: \t{}".format(created_time, modified_time, access_time))
    return stats


# Run Reports
@cache
def reportDP():
    global targetFileDP

    # Set Path & Create Dataframe
    targetPathDP = str(dataPath) + str(targetFileDP)
    dfSourceDP = pd.read_csv(targetPathDP, usecols=range(50), engine='c')

    # Define Workbook & Worksheet
    xl = xw.App(visible=False)
    wb = xl.books.open(destinationDP)
    ws = wb.sheets['ALL JOBS']

    # Start In Cell F2 & Remove Dataframe Headers/Indices
    ws.range('F2').options(index=False, header=False).value = dfSourceDP

    # Save & Close
    wb.save(destinationDP)
    wb.close()

    # Closing Response
    messagebox.showinfo("Status", "Finished.")


@cache
def reportFP():
    global targetFileFP

    # Set Path & Create Dataframe
    targetPathFP = str(dataPath) + str(targetFileFP)
    dfSourceFP = pd.read_csv(targetPathFP, usecols=range(50), engine='c')

    # Define Workbook & Worksheet
    xl = xw.App(visible=False)
    wb = xl.books.open(destinationFP)
    ws = wb.sheets['All Jobs_FP']

    # Start In Cell F2 & Remove Dataframe Headers/Indices
    ws.range('E2').options(index=False, header=False).value = dfSourceFP

    # Save & Close
    wb.save(destinationFP)
    wb.close()

    # Closing Response
    messagebox.showinfo("Status", "Finished.")


# Check Workbooks For Formula Errors
@cache
def checkDP():
    global errors
    wb = openpyxl.load_workbook(destinationDP, data_only=True)
    ws = wb['ALL JOBS']
    errors = 0
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=12, max_col=5):
        for cell in row:
            if cell.value == '#VALUE!':
                errors += 1
    resultDP = (str(errors) + " Errors Found")
    messagebox.showinfo("Status", resultDP)


@cache
def checkFP():
    wb = openpyxl.load_workbook(destinationFP, data_only=True)
    ws = wb['All Jobs_FP']
    errors = 0
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=12, max_col=5):
        for cell in row:
            if cell.value == '#VALUE!':
                errors += 1
    resultFP = (str(errors) + " Errors Found")
    messagebox.showinfo("Status", resultFP)
